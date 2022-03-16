import os
import fnmatch
import xml.etree.ElementTree as et
import openpyxl as xl



tag_filter = {'text', 'undiscoveredTooltip', 'power', 'desc', 'visitedTooltip', 'unvisitedTooltip', 'title', 'flavorType', 'short',"header","secretName","secretDescription","class"}
xls_cnt = 2
xls_len = 0
xls_it = None
def write_seg(workbook,name,node,*args):
    if node.text == None or node.text.strip() == "" or node.text.isnumeric(): return ""#exclude numbers and blanks
    if node.tag not in tag_filter: return ""   #filter by tags

    if 'name' in node.attrib: name=node.attrib['name']
    else: name=''
    workbook.append([node.text,"",name])   #["source","target","context","comments"]
    return name

def rewrite_seg(workbook,name,node,*args):
    global xls_cnt
    if node.text == None or node.text.strip() == "" or node.text.isnumeric(): return #exclude numbers and blanks
    if node.tag not in tag_filter: return   #filter by tags

    if node.text != workbook[xls_cnt][0].value:
        workbook[xls_cnt][1].value=node.text
    #for row in workbook.iter_rows(min_row=xls_cnt):
    #    #print(f"{row[0].value}: {text}")
    #    if text != row[0].value:
    #        row[1].value=text
    #        print(f"{row[1].value} <- {text} : {workbook}")
    #    break
    xls_cnt+=1

def patch_seg(workbook,name,node,*args):
    global xls_cnt
    global xls_len
    global xls_it
    if node.text == None or node.text.strip() == "" or node.text.isnumeric(): return #exclude numbers and blanks
    if node.tag not in tag_filter: return   #filter by tags

    curcell = next(xls_it)
    if node.text == curcell[0].value:     #FIXME: PERFORMANCE!!!!!
        if curcell[1].value != None:    #원문이랑 같지 않다면 패치
            #print(f"{node.text.strip()}->{workbook[xls_cnt][1].value.strip()}")
            node.text = curcell[1].value
            print(f"{xls_cnt}/{xls_len}",end='\r')

    #for row in workbook.iter_rows(min_row=xls_cnt):
    #    if text == row[2].value:
    #        if text != row[0].value:    #원문이랑 같지 않다면 패치
    #            text = row[1].value
    #        break
    xls_cnt+=1

def iter_node(func,workbook,root,segname):
    segname = func(workbook,segname,root)
    for node in root:
        #print(f"{node.tag} : {node.text}")
        segname = func(workbook,segname,node)

        for children in node:
            iter_node(func,workbook,children,segname)


def print_nodes(root):    #FOR DEBUG
    print("attrib:",root.attrib,"\ntext:",root.text,"\ntag:",root.tag,"\n")
    for node in root:
        print("attrib:",node.attrib,"\ntext:",node.text,"\ntag:",node.tag,"\n")

        for children in node:
            print_nodes(children)

def bootstrap(readdir,readpattern,writedir,func='write_seg',xlsdir=None):

    parser = et.XMLParser(target=et.TreeBuilder(insert_comments=True))  

    for file in os.listdir(f"./{readdir}"):
        if fnmatch.fnmatch(file,readpattern):
            print(f"parsing {file}...")
            try:
                if xlsdir == None:
                    xls=xl.workbook.Workbook()
                    xls.active.append(["source","target","context","comments"])
                else:
                    global xls_cnt
                    global xls_len
                    global xls_it
                    xls_cnt = 2
                    xls=xl.load_workbook(filename=f"./{xlsdir}/{file}.xlsx",read_only=True if func =='patch_seg' else False,data_only=True)
                    xls_it=xls.active.iter_rows(min_row=2)
                    xls_len = xls.active.max_row #for showing progress?

                #xml=parser.parse(f"./{readdir}/{file}")
                xml=et.parse(f"./{readdir}/{file}")
                xml_root = xml.getroot()

                #print_nodes(xml_root)
                iter_node(eval(func),xls.active,xml_root,"")

                if func == 'patch_seg':
                    xml.write(f"./{writedir}/{file}",encoding='utf-8',xml_declaration=False)  
                else:
                    if xls.active.max_row > 1:
                        xls.save(filename=f'./{writedir}/{file}.xlsx')
                    else:
                        print(f"WARNING: no texts in {file}, did not write xlsx.")

            except et.ParseError as e:
                print(f"ERROR: cannot parse {file}: {e.msg}")
                xls=None
            except FileNotFoundError as e:
                print(f"{str(e)}, skipped")

